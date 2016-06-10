# -*- coding: utf-8 -*-
def BelowGradeSimple(ExcelFile):
    """MAIN Function. It receives the Excel File called "BelowGrade.xlsx" (in python format) with all the properties of the basement
    (below grade surfaces). It will use the information to calculate the heat loss through each component of the basement, 
    and return a list of dictionaries  with all the values, including the total heat loss. It uses 2 external functions 
    (FindData and FillResults), defined separately. Additionally, it will update the excel file with the results, in
    the sheet called "Results", with the heat losses, the U values, and the Areas"""
     
    from math import log
    q_Below_Grade=[None]*4 #empty list, it will be filled with dictionaries#
    qw={} #empty dictionary#
    qf={} #empty dictionary#
    qr={} #empty dictionary#
    qt={} #empty dictionary#
    
    U=[None]*4 #empty list, it will be filled with dictionaries#
    Uw={}#empty dictionary#
    Uf={}#empty dictionary#
    U3={}#empty dictionary#
    U4={}#empty dictionary#
    
    Areas=[None]*4 #empty list, it will be filled with dictionaries#
    Aw={}#empty dictionary#
    Af={}#empty dictionary#
    A3={}#empty dictionary#
    A4={}#empty dictionary#

    basement = findData(ExcelFile) # Creation of the basement dictionary using the function "findData"

    # The following lines are assignment of basement properties to individual variables......
    insulation=basement["Is the ground floor insulated)"]
    type_roof=basement["Ground_floor_type"]
    W1=basement["Width1"]
    W2=basement["Width2"]
    Tin=basement["T_inside"]
    Tmgr=basement["T_mean_ground"] 
    A=basement["T_ground_amplitude"] 
    R=basement["Wall_resistance"] 
    R_u=basement["Uninsulation_resistance"]
    z2=basement["Depht"] 
    z1=basement["Uninsulation_height"]
    Tin_b=basement["Internal_temperature_of_the_building"]
    Tgr=Tmgr-A #minimum ground surface temperature#
    k=1.4 #soil thermal conductivity, [W/(m·K)]#
    pi=3.1416
    zf=z2 
    wb=min(W1,W2) #basement width (shortest dimension), [m]#
    #...........................................................................................

# This part of the code just reads a table from excel and determines a Fp factor.................
    Excel_Table = load_workbook("BelowGrade.xlsx")# Here we select the excel SHEET in which the table is
    Table_Data = Excel_Table.get_sheet_by_name("Table")# Here we select the sheet having the table
    Data = Table_Data.columns[0:][0:]   # Here we create a Matrix with all the info we extracted from the table 
    Table_values=[]# Here we create a vector with All the Dictionaries of each value in the table
    for h in range(1,3,1): #this numbers coinicide with the dimensions of the table
        s=1
        for j in Data[h][1:]:
            x={}
            x["Value"]=float(j.value) #here we extract the value
            x["Type"]=float(Data[0][s].value)# here we take the X reference of that value
            x["Insulation"]=(Data[h][0].value).encode('utf-8')# here we take the Y reference of that value
            s=s+1
            Table_values.append(x) #Here we update the Vector
    for t in Table_values:
        if type_roof==t["Type"]:
            if insulation==t["Insulation"]:
                Fp=t["Value"]
#........................................................................
    
    if z1 == 0: # This part of the code performs the calculations for the case of UNIFORM INSULATION all over the wall
        U_wall=2*k/(pi*(z2-z1))* ( log(z2+2*k*R/pi)-log(z1+2*k*R/pi) ) #average U-factor for wall region defined by z1 and z2, [W/(m2·K)]# From ASHRAE, chap 18, Eq 29
        U_floor=2*k/(pi*wb)* ( log(wb/2+zf/2+k*R/pi)-log(zf/2+k*R/pi) ) #average U-factor for below-grade floor [W/(m2·K)]# From ASHRAE, chap 18, Eq 30
        
        A_wall=2*(W1+W2)*(z2-z1) #total wall area#
        A_floor=W1*W2 #floor area#
        
        q_wall=U_wall*A_wall*(Tin-Tgr) #heat loss through wall#
        q_floor=U_floor*A_floor*(Tin-Tgr) #heat loss through floor#
        q_roof=2*(W1+W2)*Fp*(Tin_b-Tin) #heat loss through roof#
        
        qw["Component"]="Below Walls Loss [W]" #compiling dictionary qw with key value "Below Walls Loss [W]"#
        qw["Value"]=q_wall #compiling dictionary qw with value q_wall#
        qf["Component"]="Below Floor Loss [W]" #compiling dictionary qf with key value "Below Floor Loss [W]"#
        qf["Value"]=q_floor #compiling dictionary qf with value q_floor#
        qr["Component"]="Roof Loss [W]" #compiling dictionary qr with key value "Roof Loss [W]"#
        qr["Value"]=q_roof #compiling dictionary qr with value q_roof#
        qt["Component"]="Total Heat Loss [W]" #compiling dictionary qt with key value "Total Heat Loss [W]"#
        qt["Value"]=q_wall+q_floor+q_roof #compiling dictionary qt with value the sum of q_wall and q_floor#
        
        q_Below_Grade[0]=qw #the dictionary qw is the first element of the list q_Below_Grade#
        q_Below_Grade[1]=qf #the dictionary qf is the second element of the list q_Below_Grade#
        q_Below_Grade[2]=qr #the dictionary qt is the third element of the list q_Below_Grade#
        q_Below_Grade[3]=qt #the dictionary qt is the fourth element of the list q_Below_Grade#

  # The following part recovers the intermediate calculations of U and Areas, to fill their respective List .................     
        Uw["Component"]="U wall [w/m2.K]"
        Uw["Value"]=U_wall
        Uf["Component"]="U Floor [w/m2.k]"
        Uf["Value"]=U_floor
        U3["Component"]="N/A"
        U3["Value"]=0
        U4["Component"]="N/A"
        U4["Value"]=0
        
        Aw["Component"]="Area wall [m2]"
        Aw["Value"]=A_wall
        Af["Component"]="Area Floor [m2]"
        Af["Value"]=A_floor
        A3["Component"]="N/A"
        A3["Value"]=0
        A4["Component"]="N/A"
        A4["Value"]=0
        
        U[0]=Uw
        U[1]=Uf
        U[2]=U3
        U[3]=U4
        
        Areas[0]=Aw
        Areas[1]=Af
        Areas[2]=A3
        Areas[3]=A4
 #...........................................................................................................
        
        FillResults(q_Below_Grade,U,Areas) #calling the function "FillResults" to update the excel file with the results"
        return q_Below_Grade,Fp
     
    else: #The following part of the code performs the same calculations as before, but for the the case of PARTIAL INSULATION......  
        U_wall=2*k/(pi*(z2-z1))* ( log(z2+2*k*R/pi)-log(z1+2*k*R/pi) ) #average U-factor for wall region defined by z1 and z2, [W/(m2·K)]#From ASHRAE, chap 18, Eq 29
        A_wall=2*(W1+W2)*(z2-z1) #total insulated wall area#
        U_floor=2*k/(pi*wb)* ( log(wb/2+zf/2+k*R/pi)-log(zf/2+k*R/pi) ) #average U-factor for below-grade floor [W/(m2·K)]#From ASHRAE, chap 18, Eq 30
        A_floor=W1*W2 #floor area#

        z2=z1
        z1=0
        U_wall_partial=2*k/(pi*(z2-z1))* ( log(z2+2*k*R_u/pi)-log(z1+2*k*R_u/pi) ) #average U-factor for uninsulated wall region defined by 0 and z1, [W/(m2·K)]#
        A_wall_partial=2*(W1+W2)*z2 #total uninsulated wall area#
        
        U_wall_tot=(U_wall*A_wall+U_wall_partial*A_wall_partial)/(A_wall+A_wall_partial) #weighted average U-factor for insulated and uninsulated wall region, [W/(m2·K)]#
        
        q_wall=U_wall_tot*(A_wall+A_wall_partial)*(Tin-Tgr) #heat loss through wall#
        q_floor=U_floor*A_floor*(Tin-Tgr) #heat loss through floor#
        q_roof=2*(W1+W2)*Fp*(Tin_b-Tin) #heat loss through roof#
        
        qw["Component"]="Below Walls Loss [W]" #compiling dictionary qw with key value "Below Walls Loss [W]"#
        qw["Value"]=q_wall #compiling dictionary qw with value q_wall#
        qf["Component"]="Below Floor Loss [W]" #compiling dictionary qf with key value "Below Floor Loss [W]"#
        qf["Value"]=q_floor #compiling dictionary qf with value q_floor#
        qr["Component"]="Roof Loss [W]" #compiling dictionary qr with key value "Roof Loss [W]"#
        qr["Value"]=q_roof #compiling dictionary qr with value q_roof#
        qt["Component"]="Total Heat Loss [W]" #compiling dictionary qt with key value "Total Heat Loss [W]"#
        qt["Value"]=q_wall+q_floor+q_roof #compiling dictionary qt with value the sum of q_wall and q_floor#

    # Recovering of intermediate calculation to fill their respective lists..................
        Uw["Component"]="U wall [w/m2.K]"
        Uw["Value"]=U_wall_tot
        Uf["Component"]="U Floor [w/m2.K]"
        Uf["Value"]=U_floor
        U3["Component"]="N/A"
        U3["Value"]=0
        U4["Component"]="N/A"
        U4["Value"]=0
        
        Aw["Component"]="Area wall [m2]"
        Aw["Value"]=A_wall+A_wall_partial
        Af["Component"]="Area Floor [m2]"
        Af["Value"]=A_floor
        A3["Component"]="N/A"
        A3["Value"]=0
        A4["Component"]="N/A"
        A4["Value"]=0
        
        U[0]=Uw
        U[1]=Uf
        U[2]=U3
        U[3]=U4
        
        Areas[0]=Aw
        Areas[1]=Af
        Areas[2]=A3
        Areas[3]=A4
    #.......................................................................
        q_Below_Grade[0]=qw #the dictionary qw is the first element of the list q_Below_Grade#
        q_Below_Grade[1]=qf #the dictionary qf is the second element of the list q_Below_Grade#
        q_Below_Grade[2]=qr #the dictionary qt is the third element of the list q_Below_Grade#
        q_Below_Grade[3]=qt #the dictionary qt is the fourth element of the list q_Below_Grade#
        
        FillResults(q_Below_Grade,U,Areas) #calling the function "FillResults" to update the excel file"
        return q_Below_Grade,Fp
       


def findData(ExcelFile):
    """This function find the data of below-grade surfaces in the Excel File BelowGrade.xlsx (must be kept in same folder)
    and returns dictionary holding all the data"""
    WindowData = ExcelFile.get_sheet_by_name("Data")# Here we extract the data from the sheet called "Data"
    basement={}#we create a dictionary for saving all the properties of the basement
    basement["Depht"] = float(WindowData.columns[1][1].value) #depths of bottom of wall segment under consideration, [m]#
    basement["Wall_resistance"] = float(WindowData.columns[1][2].value) #total resistance of wall, insulation, and indoor surface resistance,[(m2·K)/W]#
    basement["Width1"] = float(WindowData.columns[1][3].value)# Basement width 1
    basement["Width2"] = float(WindowData.columns[1][4].value)# Basement width 2
    basement["T_inside"] = float(WindowData.columns[1][5].value) #below-grade space air temperature, [°C]#
    basement["T_mean_ground"] = float(WindowData.columns[1][6].value) #mean ground temperature, [°C], estimated from the annual average air temperature or from well-water temperatures#
    basement["T_ground_amplitude"] = float(WindowData.columns[1][7].value) #ground surface temperature amplitude, [°C]#
    basement["Uninsulation_height"] = float(WindowData.columns[1][8].value) # partial insulation height
    basement["Uninsulation_resistance"] = float(WindowData.columns[1][9].value)# partial insulation resistance
    basement["Ground_floor_type"] = float(WindowData.columns[1][10].value)# the material of the "roof" of the basement
    basement["Is the ground floor insulated)"] =(WindowData.columns[1][11].value)# boolean for wether the "roof" is insulated or not
    basement["Internal_temperature_of_the_building"] = float(WindowData.columns[1][12].value)# inside temperature of the building.
    return basement





def FillResults(q_Below_Grade,U,Areas):
    """This function receives 3 list of dictionaries (Heat Losses, U values, and Areas ), each dictionary belongs to a component of their respective list.
    It will take the lists and traspass the information into a table in excel, in the sheet called Results"""

    from openpyxl import *
    ExcelFile = load_workbook("BelowGrade.xlsx") #loading the excel file#
    WindowData = ExcelFile.get_sheet_by_name("Results")
    for j in range(0,8): #"for loop" for columns#
      for h in range(0,4): #"for loop" for rows#
        if j==0:
            WindowData.columns[j][h].value = q_Below_Grade[h]["Component"] #filling the first column of the excel sheet with the key value of the dictionaries for heat losses#
            ExcelFile.save("BelowGrade.xlsx")
            
        elif j==1:
            WindowData.columns[j][h].value = q_Below_Grade[h]["Value"] #filling the second column of the excel sheet with the values of the dictionaries for heat losses#
            ExcelFile.save("BelowGrade.xlsx")

        elif j==3:
           WindowData.columns[j][h].value = U[h]["Component"] #filling the fourth column of the excel sheet with the key values of the U-factor#
           ExcelFile.save("BelowGrade.xlsx")
                      
        elif j==4:
           WindowData.columns[j][h].value = U[h]["Value"] #filling the fifth column of the excel sheet with the values of the dictionaries for U-factor#
           ExcelFile.save("BelowGrade.xlsx")
           
        elif j==6:
           WindowData.columns[j][h].value = Areas[h]["Component"] #filling the seventh column of the excel sheet with the key values of the areas#
           ExcelFile.save("BelowGrade.xlsx")
           
        elif j==7:
           WindowData.columns[j][h].value = Areas[h]["Value"] #filling the eighth column of the excel sheet with the values of the dictionaries for areas#
           ExcelFile.save("BelowGrade.xlsx")
           
    return "Results have been Uploaded"


from openpyxl import *
ExcelFile = load_workbook("BelowGrade.xlsx")
qbg=BelowGradeSimple(ExcelFile)
print qbg
